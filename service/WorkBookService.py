# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import logging


class WorkBookService:
    def __init__(self, filename):
        self.wb = load_workbook(filename)
        self.logger = logging.getLogger("WorkBookService")

    # 获取列名
    def get_column_names(self):
        ws = self.wb.active
        column_names = []
        for i in range(1, ws.max_column+1):
            cell = get_column_letter(i) + '1'
            column_names.append(ws[cell].value)
        self.logger.info("column names: {0}".format(column_names))
        return column_names

    # 获取用例
    def get_test_cases(self):
        ws = self.wb.active
        row_nums = ws.max_row + 1

        self.logger.info("用例最大行数：{0}".format(row_nums))

        title_column_letter = get_column_letter(self.get_column_names().index('Name') + 1)
        step_column_letter = get_column_letter(self.get_column_names().index('STEPS') + 1)
        result_column_letter = get_column_letter(self.get_column_names().index('Result') + 1)
        data_column_letter = get_column_letter(self.get_column_names().index('Test Data') + 1)

        cases = []
        for r in range(2, row_nums):
            step_cell = step_column_letter + str(r)
            result_cell = result_column_letter + str(r)
            data_cell = data_column_letter + str(r)
            title_cell = title_column_letter + str(r)
            step = {'step': ws[step_cell].value, 'data': ws[data_cell].value, 'result': ws[result_cell].value}

            if ws[title_cell].value:
                case = dict()
                case['title'] = ws[title_cell].value
                case['steps'] = [step]
                cases.append(case)
            else:
                cases[-1]['steps'].append(step)

        return cases


if __name__ == '__main__':
    wbs = WorkBookService(r"")
    wbs.get_column_names()
    wbs.get_test_cases()
